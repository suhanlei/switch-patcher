"""
Excel读写模块
- 读取设备清单（适配h3c_hosts格式，支持多Sheet选择）
- 逐阶段回写执行状态到Excel（线程安全，带文件锁）
- 计算本地补丁文件MD5
"""

import hashlib
import threading
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

# 文件锁，防止多线程并发写Excel时数据错乱
LOCK = threading.Lock()

# 内部列名 → Excel表头的映射
COLUMN_MAP = {
    "hostname": "Hostname",
    "mgmt_ip": "Mgmt_IP",
    "vendor": "Vendor",
    "patch_now": "patch_now",
    "patch_new": "patch_new",
    "patch_file": "patch_file",
    "md5_base": "patch1_md5_base",
    "md5_uploaded": "patch1_md5_uploaded",
    "scp_status": "scp_status",
    "upload_success": "upload_success",
    "update_result": "update_result",
}


@dataclass
class DeviceInfo:
    """从Excel读取的单台设备信息"""
    hostname: str
    mgmt_ip: str
    vendor: str
    patch_file: str
    row_index: int          # 在Excel中的行号（1-based），用于回写定位
    excel_path: str = ""    # Excel文件路径（回写时需要，由batch_engine注入）
    sheet_name: str = ""    # Excel Sheet名称（回写时需要，由batch_engine注入）
    patch_now: str = ""     # 当前补丁版本（工具回写）
    patch_new: str = ""     # 目标补丁版本（工具回写）
    md5_base: str = ""      # 本地补丁文件MD5（工具回写）
    md5_uploaded: str = ""   # 设备端文件MD5（工具回写）
    scp_status: str = ""    # 设备连接与服务状态（工具回写）：
                            #   scp/sftp/scp_sftp = SCP/SFTP已开启（隐含登录OK）
                            #   none = 未开启（隐含登录OK）
                            #   unreachable = 设备不可达
                            #   login_fail = SSH登录失败
    upload_success: str = ""  # 上传状态（工具回写）
    update_result: str = ""   # 升级结果（工具回写）


@dataclass
class DeviceResult:
    """单台设备执行结果，用于汇总报告"""
    hostname: str
    mgmt_ip: str
    vendor: str
    status: str = "pending"     # success / partial / failed / skipped
    pre_check_ok: bool = False  # 预检查是否通过
    transfer_ok: bool = False   # 文件传输是否成功
    patch_applied: bool = False  # 补丁是否已激活
    post_check_ok: bool = False  # 后检查是否通过
    commands_total: int = 0     # 总命令数
    commands_applied: int = 0   # 成功执行命令数
    commands_failed: int = 0    # 失败命令数
    cpu_before: float | None = None   # 补丁前CPU使用率
    cpu_after: float | None = None    # 补丁后CPU使用率
    mem_before: float | None = None   # 补丁前内存使用率
    mem_after: float | None = None    # 补丁后内存使用率
    patch_now: str = ""         # 当前补丁版本
    patch_new: str = ""         # 目标补丁版本
    error_message: str = ""     # 错误信息
    start_time: datetime | None = None  # 开始时间
    end_time: datetime | None = None    # 结束时间


def list_sheets(excel_path: str) -> list[str]:
    """列出Excel中所有Sheet名称，用于灰度选择"""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    names = wb.sheetnames[:]
    wb.close()
    return names


def read_devices(excel_path: str, sheet_name: str | None = None) -> list[DeviceInfo]:
    """
    从Excel读取设备清单
    - sheet_name: 指定Sheet名称（灰度补丁），为None时使用活动Sheet
    - 返回: DeviceInfo列表
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 解析表头，建立列名到列号的映射
    headers = [cell.value for cell in ws[1]]
    col_indices = {}
    for idx, h in enumerate(headers, 1):
        if h and h in COLUMN_MAP.values():
            col_indices[h] = idx

    # 记住实际使用的Sheet名称，供回写时定位
    actual_sheet = sheet_name or ws.title

    devices = []
    for row_idx in range(2, ws.max_row + 1):
        hostname = ws.cell(row=row_idx, column=col_indices.get("Hostname", 1)).value
        mgmt_ip = ws.cell(row=row_idx, column=col_indices.get("Mgmt_IP", 2)).value
        vendor = ws.cell(row=row_idx, column=col_indices.get("Vendor", 3)).value
        patch_file = ws.cell(row=row_idx, column=col_indices.get("patch_file", 6)).value
        patch_now = ws.cell(row=row_idx, column=col_indices.get("patch_now", 4)).value
        upload_success = ws.cell(row=row_idx, column=col_indices.get("upload_success", 10)).value
        update_result = ws.cell(row=row_idx, column=col_indices.get("update_result", 11)).value
        scp_status = ws.cell(row=row_idx, column=col_indices.get("scp_status", 9)).value

        # 跳过空行
        if not hostname or not mgmt_ip:
            continue

        devices.append(DeviceInfo(
            hostname=str(hostname).strip(),
            mgmt_ip=str(mgmt_ip).strip(),
            vendor=str(vendor).strip() if vendor else "",
            patch_file=str(patch_file).strip() if patch_file else "",
            row_index=row_idx,
            patch_now=str(patch_now).strip() if patch_now else "",
            upload_success=str(upload_success).strip() if upload_success else "",
            update_result=str(update_result).strip() if update_result else "",
            scp_status=str(scp_status).strip() if scp_status else "",
            sheet_name=actual_sheet,
        ))

    wb.close()
    return devices


def write_cell(excel_path: str, row_index: int, col_name: str, value: str, sheet_name: str | None = None):
    """
    回写单个单元格到Excel（线程安全）
    - row_index: Excel行号
    - col_name: 内部列名（如"scp_status"）
    - value: 要写入的值
    - sheet_name: 目标Sheet名称，为None时使用活动Sheet
    """
    col_key = COLUMN_MAP.get(col_name)
    if not col_key:
        return

    with LOCK:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        # 查找列号，若列不存在则自动追加
        headers = [cell.value for cell in ws[1]]
        col_idx = None
        for idx, h in enumerate(headers, 1):
            if h == col_key:
                col_idx = idx
                break

        if col_idx is None:
            col_idx = len(headers) + 1
            ws.cell(row=1, column=col_idx, value=col_key)

        ws.cell(row=row_index, column=col_idx, value=value)
        wb.save(excel_path)
        wb.close()


def calc_md5(file_path: str) -> str:
    """计算本地文件的MD5哈希值，用于与设备端校验比对"""
    h = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()
